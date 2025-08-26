from src import mongo
from genericpath import exists
from mongoengine import Document, StringField, EmbeddedDocument, EmbeddedDocumentField, ListField, DictField
import openpyxl
from pymongo import MongoClient
from bson import ObjectId
import json
from io import BytesIO
import os

# Connexion directe MongoDB en cas de problème avec mongo.get_db
try:
    db = mongo.get_db
    if db is None:
        raise Exception("mongo.get_db est None")
except:
    print("⚠️  Utilisation de la connexion MongoDB directe")
    client = MongoClient('mongodb://localhost:27017/')
    db = client['cac_perform']


# ==============================
#  Clients
# ==============================
class Client(Document):

    @classmethod
    def afficher_clients(cls):
        try:
            cust_data = list(db.Client.find().sort({"_id": -1}))
            for client in cust_data:
                client['_id'] = str(client['_id'])
            return cust_data
        except Exception as e:
            print(f"An exception : {str(e)}")

    @classmethod
    def afficher_missions(cls, id):
        results = []
        final = []
        query = {"id_client": id}
        results = db.Mission1.find(query)
        for result in results:
            result['_id'] = str(result['_id'])
            final.append(result)
        return final

    @classmethod
    def ajouter_client(cls, data):
        try:
            cust_data = db.Client.insert_one(data)
            inserted_id = str(cust_data.inserted_id)
            return inserted_id
        except Exception as e:
            print(f'An exception occurred: {str(e)}')

    @classmethod
    def modifier_client(cls, data):
        try:
            _id = str(data['_id'])
            id_client = ObjectId(_id)

            nom = data['nom']
            activite = data['activite']
            referentiel = data['referentiel']
            forme_juridique = data['forme_juridique']
            capital = data['capital']
            siege_social = data['siege_social']
            adresse = data['adresse']
            n_cc = data['n_cc']

            if id_client:
                db.Client.update_one(
                    {"_id": id_client},
                    {"$set": {
                        "nom": nom,
                        "activite": activite,
                        "referentiel": referentiel,
                        "forme_juridique": forme_juridique,
                        "capital": capital,
                        "siege_social": siege_social,
                        "adresse": adresse,
                        "n_cc": n_cc,
                    }}
                )
                return "Updated succeeded"
        except Exception as e:
            print(f'An exception occurred: {str(e)}')

    @classmethod
    def info_client(cls, id_clit):
        try:
            id_client = id_clit
            info = db.Client.find_one({"_id": id_client})

            valeurs = {}
            if info:
                valeurs = {
                    "_id": str(id_client),
                    "nom": info['nom'],
                    "activite": info['activite'],
                    "siege_social": info['siege_social'],
                    "adresse": info['adresse'],
                    "referentiel": info['referentiel'],
                    "forme_juridique": info['forme_juridique'],
                    "capital": info['capital'],
                    "n_cc": info['n_cc'],
                }
            return valeurs
        except Exception as e:
            print(f"An exception : {str(e)}")
            return None

    def choix_referentiel(self):
        try:
            ref = db.Grouping.find()
            return ref
        except Exception as e:
            print(f"An exception as occured: {str(e)}")


# ==============================
#  Manager
# ==============================
class Manager(Document):
    _id = StringField()
    email = StringField(required=True)
    mot_de_passe = StringField(required=True)

    def sign_in(self, data):
        try:
            email = data['mail']
            mdp = data['pwd']
            if db.Manager.find_one({"email": email, "mot_de_passe": mdp}):
                return "success"
        except Exception as e:
            print(f"An exception occurred : {str(e)}")
            return None


# ==============================
#  Modèles embarqués
# ==============================
class LigneComptable(EmbeddedDocument):
    num_compte = StringField(required=True, unique=True)
    libelle = StringField()
    solde = DictField()


class BalanceComptable(EmbeddedDocument):
    lignes = ListField(EmbeddedDocumentField(LigneComptable))


# ==============================
#  Mission
# ==============================
class Mission(Document):
    # champs (non utilisés par mongoengine ici mais conservés)
    balance = EmbeddedDocumentField(BalanceComptable)
    annee_auditee = ListField()

    # ---------- Revue analytique ----------
    def revue_analytique(self, id_mission):
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        variations = mission.get("balance_variation", [])

        # mapping_efi.json (un niveau au-dessus de ce fichier)
        mapping_path = os.path.join(os.path.dirname(__file__), "..", "mapping_efi.json")
        with open(mapping_path, "r", encoding="utf-8") as f:
            mapping_root = json.load(f)
            mapping = mapping_root["structure"]

        # seuils (matérialité choisie si existe)
        materiality = mission.get("materiality", [])
        choice = next((m for m in materiality if m.get("choice")), None)
        perf_mat = abs(int(choice.get("performance_materiality", 0))) if choice else 0

        def collect_prefixes(bloc):
            prefixes = []
            for key in ("brut_cpt", "amor_cpt", "net_cpt", "brut_except_cpt", "amor_except_cpt", "net_except_cpt"):
                if key in bloc and bloc[key]:
                    if isinstance(bloc[key], str):
                        prefixes += [p.strip() for p in bloc[key].split(",") if p.strip()]
                    elif isinstance(bloc[key], list):
                        prefixes += [str(p).strip() for p in bloc[key] if str(p).strip()]
            return prefixes

        def map_efi(numero):
            refs = []
            for bloc in mapping:
                prefixes = collect_prefixes(bloc)
                if any(str(numero).startswith(p) for p in prefixes):
                    refs.append(bloc["ref"])
            return sorted(list(set(refs)))

        out = []
        for row in variations:
            n = int(row.get("solde_n", 0) or 0)
            n1 = int(row.get("solde_n1", 0) or 0)
            delta = n - n1
            delta_pct = (delta / (abs(n1) if n1 else 1.0))
            efi_refs = map_efi(row.get("numero_compte", ""))

            # Commentaire automatique (EFI + Matérialité)
            commentaire_auto = (
                f"EFI: {', '.join(efi_refs) if efi_refs else 'Aucun'} — "
                f"Matérialité: {'Oui' if (perf_mat and abs(delta) >= perf_mat) else 'Non'}"
            )

            out.append({
                "numero_compte": row.get("numero_compte"),
                "libelle": row.get("libelle"),
                "solde_n": n,
                "solde_n1": n1,
                "delta_abs": delta,
                "delta_pct": delta_pct,
                "commentaire_auto": commentaire_auto,
                "commentaire_perso": ""  # Commentaire personnalisé vide par défaut
            })

        db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"revue_analytique": out}})
        return out

    def update_commentaire_perso(self, id_mission, numero_compte, commentaire_perso):
        """
        Met à jour le commentaire personnalisé pour un compte spécifique
        """
        try:
            print(f"Début de mise à jour du commentaire pour le compte {numero_compte}")
            
            # Vérifier d'abord que la mission existe
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            if not mission:
                print(f"Mission {id_mission} non trouvée")
                return False
            
            # Vérifier que la revue analytique existe
            if "revue_analytique" not in mission:
                print(f"Revue analytique non trouvée pour la mission {id_mission}")
                return False
            
            # Trouver l'index du compte dans la revue analytique
            revue_analytique = mission["revue_analytique"]
            compte_index = None
            
            for i, item in enumerate(revue_analytique):
                if item.get("numero_compte") == numero_compte:
                    compte_index = i
                    break
            
            if compte_index is None:
                print(f"Compte {numero_compte} non trouvé dans la revue analytique")
                return False
            
            # Vérifier si le commentaire a réellement changé
            current_comment = revue_analytique[compte_index].get("commentaire_perso", "")
            if current_comment == commentaire_perso:
                print(f"Commentaire identique pour le compte {numero_compte}, aucune modification nécessaire")
                return True  # Retourner True car c'est un succès (pas d'erreur)
            
            print(f"Mise à jour du commentaire: '{current_comment}' -> '{commentaire_perso}'")
            
            # Mettre à jour le commentaire personnalisé
            result = db.Mission1.update_one(
                {"_id": ObjectId(id_mission)},
                {"$set": {f"revue_analytique.{compte_index}.commentaire_perso": commentaire_perso}}
            )
            
            success = result.modified_count > 0
            if success:
                print(f"Commentaire mis à jour avec succès pour le compte {numero_compte}")
            else:
                print(f"Aucune modification effectuée pour le compte {numero_compte} (modified_count: {result.modified_count})")
            
            return success
            
        except Exception as e:
            print(f"Erreur lors de la mise à jour du commentaire: {e}")
            import traceback
            traceback.print_exc()
            return False

    # ---------- Nouvelle mission ----------
    def nouvelle_mission(self, balances, annee_auditee, id_client, date_debut, date_fin):
        balance_ids = []
        les_balance_n_n1 = []
        annee_balance = annee_auditee

        for balance in balances:
            balance_created = self.creation_balance(balance, int(annee_auditee), id_client)
            annee_auditee = int(annee_auditee) - 1

            tuple_en_tableau = list(*[balance_created])
            balance_ids.append(tuple_en_tableau[0])
            les_balance_n_n1.append(tuple_en_tableau[1])

        balance_variation = self.rapprochement_des_balances(
            les_balance_n_n1[0], les_balance_n_n1[1]
        )

        grouping_model = self.create_grouping(balance_variation)
        etats = self.prod_efi(les_balance_n_n1[0], les_balance_n_n1[1], balance_variation)

        result = db.Mission1.insert_one({
            "id_client": id_client,
            "annee_auditee": str(annee_balance),
            "date_debut": date_debut,
            "date_fin": date_fin,
            "balances": balance_ids,
            "balance_variation": balance_variation,
            "grouping": grouping_model,
            "efi": etats,
            "materiality": []
        })

        insert_id = str(result.inserted_id)
        res = {
            "id_client": id_client,
            "annee_auditee": str(annee_balance),
            "date_debut": date_debut,
            "date_fin": date_fin,
            "balances": balance_ids,
            "balance_variation": balance_variation,
            "grouping": grouping_model,
            "efi": etats,
            "materiality": []
        }

        format_id = {"_id": insert_id, "mission": res}
        self.audit_trail(format_id['_id'])
        return format_id

    # ---------- Lecture d'une balance Excel -> Mongo (Excel uniquement) ----------
    def creation_balance(self, balance_data, annee_auditee, id_client):
        balance = balance_data
        data = []

        try:
            # Traitement des fichiers Excel uniquement
            workbook = openpyxl.load_workbook(balance)
            
            # Détection automatique de la feuille de balance
            sheet = None
            sheet_name = None
            
            # Noms de feuilles acceptés (par ordre de priorité)
            accepted_sheet_names = [
                'Balance_des_comptes',  # Nom standard
                'BALANCE_2023',         # Votre format
                'BALANCE__2024',        # Votre format
                'Sheet1',               # Format générique
                'Balance',              # Format court
                'Comptes',              # Format alternatif
            ]
            
            # Chercher la première feuille disponible
            for name in accepted_sheet_names:
                if name in workbook.sheetnames:
                    sheet = workbook[name]
                    sheet_name = name
                    print(f"✅ Feuille trouvée: '{name}' dans {balance.filename if hasattr(balance, 'filename') else 'fichier'}")
                    break
            
            # Si aucune feuille standard n'est trouvée, prendre la première
            if sheet is None and workbook.sheetnames:
                sheet = workbook[workbook.sheetnames[0]]
                sheet_name = workbook.sheetnames[0]
                print(f"⚠️  Feuille non standard utilisée: '{sheet_name}' dans {balance.filename if hasattr(balance, 'filename') else 'fichier'}")
            
            if sheet is None:
                raise Exception("Aucune feuille trouvée dans le fichier Excel")

            print(f"📊 Traitement de la feuille '{sheet_name}' avec {sheet.max_row} lignes et {sheet.max_column} colonnes")
            
            # Détection automatique de la ligne d'en-tête
            header_row = 1
            for row_idx in range(1, min(5, sheet.max_row + 1)):
                row_data = [sheet.cell(row=row_idx, column=col).value for col in range(1, min(8, sheet.max_column + 1))]
                if any('compte' in str(cell).lower() for cell in row_data if cell):
                    header_row = row_idx
                    print(f"📝 En-tête détecté à la ligne {row_idx}: {row_data}")
                    break
            
            # Traitement des lignes de données
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                if row[0] is None and row[1] is None:
                    break
                
                # Vérifier que la ligne contient des données valides
                if not row[0] or str(row[0]).strip() == '':
                    continue
                
                ligne = {}
                ligne['numero_compte'] = str(row[0])
                ligne['libelle'] = row[1] or ''
                ligne['debit_initial'] = row[2] or 0
                ligne['credit_initial'] = row[3] or 0
                ligne['debit_mvt'] = row[4] or 0
                ligne['credit_mvt'] = row[5] or 0
                ligne['debit_fin'] = int(row[6] or 0)
                ligne['credit_fin'] = int(row[7] or 0)
                ligne['solde_reel'] = ligne['debit_fin'] - ligne['credit_fin']
                ligne['solde'] = abs(ligne['solde_reel'])
                ligne['sign_solde'] = "D" if ligne['debit_fin'] >= ligne['credit_fin'] else "C"
                data.append(ligne)
            
            print(f"✅ {len(data)} lignes de données extraites de la feuille '{sheet_name}'")
                    
        except Exception as e:
            print(f"Erreur lors du traitement du fichier Excel: {str(e)}")
            raise e

        result = db.Balance.insert_one(
            {"id_client": id_client, "annee_balance": annee_auditee, "balance": data}
        )

        inserted_id = str(result.inserted_id)
        return inserted_id, data

    # ---------- Grouping ----------
    def create_grouping(self, balances_rapprochee, referentiel="syscohada"):
        grouping_path = os.path.join(os.path.dirname(__file__), "..", "grouping.json")
        with open(grouping_path, 'r', encoding='utf-8') as file:
            result = json.load(file)
        table_grouping = result[referentiel]

        for group in table_grouping:
            nber_group = group['compte']
            group['solde_n'] = sum(
                item['solde_n'] for item in balances_rapprochee
                if str(item['numero_compte']).startswith(nber_group)
            )
            group['solde_n1'] = sum(
                item['solde_n1'] for item in balances_rapprochee
                if str(item['numero_compte']).startswith(nber_group)
            )
            group['variation'] = group['solde_n'] - group['solde_n1']

            if group['variation'] == 0:
                group['variation_percent'] = 0
            elif group['solde_n1'] == 0:
                group['variation_percent'] = 100
            else:
                group['variation_percent'] = (group['variation'] / group['solde_n1']) * 100

        return table_grouping

    # ---------- Variation N vs N-1 ----------
    def rapprochement_des_balances(self, balance_n, balance_n1):
        variation_des_balances = []
        idx_n1 = {str(item.get('numero_compte')): item for item in balance_n1}

        for bal in balance_n:
            numero = str(bal['numero_compte'])
            ligne = {
                'numero_compte': numero,
                'libelle': bal['libelle'],
                'solde_n': bal['solde_reel'],
                'solde_n1': idx_n1.get(numero, {}).get('solde_reel', 0)
            }
            variation_des_balances.append(ligne)

        return variation_des_balances

    # ---------- Charges ----------
    def total_charges(self, id_mission):
        balance = db.Mission1.find_one({"_id": ObjectId(id_mission)})['balance_variation']
        charges = sum(item['solde_n'] for item in balance if str(item['numero_compte']).startswith('6'))
        return abs(charges)

    # ---------- Benchmarks ----------
    def get_benchmarks(self, id_mission):
        try:
            mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
            efi = mission['efi']

            bench = {}
            bench['total_assets'] = int(efi['actif'][-1]['net_solde_n'])
            bench['profit_before_tax'] = int(efi['pnl'][33]['net_solde_n']) + int(efi['pnl'][38]['net_solde_n'])
            bench['revenue'] = int(efi['pnl'][7]['net_solde_n'])
            bench['ebitda'] = int(efi['pnl'][23]['net_solde_n'])
            bench['expenses'] = self.total_charges(id_mission)
            return bench
        except Exception as e:
            print(f"An error there: {str(e)}")
            return None

    # ---------- Materiality ----------
    def save_materiality(self, id_mission, materialities):
        query = {"_id": ObjectId(id_mission)}
        verify = db.Mission1.find_one(query)

        if len(verify['materiality']) != 0 and materialities['benchmark'] in [item['benchmark'] for item in verify['materiality']]:
            return 0
        else:
            update = db.Mission1.update_one(query, {"$push": {"materiality": materialities}})
            return update.modified_count

    def validate_materiality(self, id_mission, benchmark):
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        materialities = mission['materiality']

        exist_choice = next((item for item in materialities if item.get('choice')), None)
        if not exist_choice:
            for item in materialities:
                if item['benchmark'] == benchmark:
                    item['choice'] = True
        else:
            exist_choice['choice'] = ""
            for item in materialities:
                if item['benchmark'] == benchmark:
                    item['choice'] = True

        result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"materiality": materialities}})
        return result.modified_count

    def get_materiality(self, id_mission):
        materiality = db.Mission1.find_one({"_id": ObjectId(id_mission)}, {"_id": 0, "materiality": 1})
        return materiality

    # ---------- Analyses grouping ----------
    def make_quantitative_analysis(self, id_mission):
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        grouping = mission['grouping']
        materiality = next((mat for mat in mission['materiality'] if mat.get('choice')), None)

        if materiality is not None:
            for item in grouping:
                value = False
                if int(item['solde_n']) >= int(materiality['materiality']):
                    value = True
                item['materiality'] = value

            result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})
            return result.modified_count
        else:
            return 0

    def make_qualitative_analysis(self, id_mission, significant):
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        grouping = mission['grouping']

        for item in grouping:
            value = next((elt['significant'] for elt in significant if elt.get('compte') == item.get('compte')), None)
            if value is None:
                item['significant'] = False
            else:
                item['significant'] = value

        result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})
        return result.modified_count

    def make_final_sm(self, id_mission):
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        grouping = mission['grouping']

        for item in grouping:
            mat = item.get('materiality', None)
            sign = item.get('significant', None)

            if (mat is not None) and (sign is not None):
                value = ""
                if mat is False and sign is True:
                    value = "non matériel significatif"
                elif mat is False and sign is False:
                    value = "non matériel non significatif"
                elif mat is True and sign is True:
                    value = "matériel significatif"
                elif mat is True and sign is False:
                    value = "matériel non significatif"
                else:
                    value = None
                item['mat_sign'] = value

        result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})
        return result.modified_count, grouping

    # ---------- Récup mission ----------
    def afficher_informations_missions(self, id_client):
        _id = id_client
        query = db.Mission1.find_one({"_id": _id})
        query['_id'] = str(query['_id'])
        return query

    # ---------- Production COMMENTAIRE ----------
    def prod_efi(self, balance_n, balance_n1, balance_variation):
        mapping_path = os.path.join(os.path.dirname(__file__), "..", "mapping_efi.json")
        with open(mapping_path, 'r', encoding='utf-8') as file:
            result = json.load(file)
        mapping = result['structure']

        # nettoyer les champs *_cpt en listes
        for mapp in mapping:
            mapp['brut_cpt'] = mapp['brut_cpt'].split(',') if mapp.get('brut_cpt') is not None else mapp.get('brut_cpt')
            mapp['amor_cpt'] = mapp['amor_cpt'].split(',') if mapp.get('amor_cpt') is not None else mapp.get('amor_cpt')
            mapp['net_cpt'] = mapp['net_cpt'].split(',') if mapp.get('net_cpt') is not None else mapp.get('net_cpt')
            mapp['brut_except_cpt'] = mapp['brut_except_cpt'].split(',') if mapp.get('brut_except_cpt') is not None else mapp.get('brut_except_cpt', [])
            mapp['amor_except_cpt'] = mapp['amor_except_cpt'].split(',') if mapp.get('amor_except_cpt') is not None else mapp.get('amor_except_cpt', [])
            mapp['net_except_cpt'] = mapp['net_except_cpt'].split(',') if mapp.get('net_except_cpt') is not None else mapp.get('net_except_cpt', [])

        datum = {}
        list_efi = ['actif', 'passif', 'pnl']
        for efi in list_efi:
            structure = []
            select_mapping = (elt for elt in mapping if elt['nature'] == efi)
            for data in select_mapping:
                row = {}
                if data.get('brut_cpt') and data.get('amor_cpt'):
                    # N
                    brut_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['brut_cpt']))
                    amor_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['amor_cpt']))

                    brut_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('brut_except_cpt', [])))
                    amor_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('amor_except_cpt', [])))

                    data['brut_solde_n'] = brut_solde_n + brut_except_n
                    data['amor_solde_n'] = amor_solde_n + amor_except_n
                    data['net_solde_n'] = data['brut_solde_n'] + data['amor_solde_n']

                    # N-1
                    brut_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['brut_cpt']))
                    amor_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['amor_cpt']))
                    net_except_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))
                    data['net_solde_n1'] = brut_n1 + amor_n1 + net_except_n1
                else:
                    net_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['net_cpt']))
                    net_solde_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['net_cpt']))

                    net_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))
                    net_except_n1_bis = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))

                    data['net_solde_n'] = net_solde_n + net_except_n
                    data['net_solde_n1'] = net_solde_n1 + net_except_n1_bis

                row['ref'] = data['ref']
                row['libelle'] = data['libelle']
                row['compte_to_be_used'] = str(data.get('brut_cpt')) + str(data.get('amor_cpt')) + str(data.get('net_cpt')) + str(data.get('brut_except_cpt')) + str(data.get('amor_except_cpt')) + str(data.get('net_except_cpt'))
                row['compte_to_be_used'] = row['compte_to_be_used'].replace('None', '')

                one = data.get('brut_cpt', []) or []
                two = data.get('amor_cpt', []) or []
                three = data.get('net_cpt', []) or []
                four = data.get('brut_except_cpt', []) or []
                five = data.get('amor_except_cpt', []) or []
                six = data.get('net_except_cpt', []) or []

                row['compte_to_be_used_off'] = list(set(one + two + three + four + five + six))

                row['brut_solde_n'] = data.get('brut_solde_n')
                row['amor_solde_n'] = data.get('amor_solde_n')
                row['net_solde_n'] = data.get('net_solde_n')
                row['net_solde_n1'] = data.get('net_solde_n1')

                structure.append(row)

            datum[efi] = structure

        return datum

    # ---------- Piste d'audit ----------
    def audit_trail(self, id_mission):
        # Créer un fichier Excel pour la piste d'audit
        wb = openpyxl.Workbook()
        sheet = wb.active

        columns = ['A', 'B', 'C', 'D', 'E']
        headers = ['Numéro compte', 'Solde n', 'Solde n-1', 'Grouping', 'Code COMMENTAIRE']

        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        balances = mission['balance_variation']
        actif = mission['efi']['actif']
        passif = mission['efi']['passif']
        pnl = mission['efi']['pnl']

        efi = actif + passif + pnl

        for i in range(len(columns)):
            sheet[columns[i] + '1'] = headers[i]

        for iteration, data in enumerate(balances):
            new_iteration = str(iteration + 2)
            sheet["A" + new_iteration] = data.get("numero_compte")
            sheet["B" + new_iteration] = data.get("solde_n")
            sheet["C" + new_iteration] = data.get("solde_n1")
            sheet["D" + new_iteration] = data.get("numero_compte")[0:2]

            list_code_efi = []
            for obj in efi:
                for elt in obj['compte_to_be_used_off']:
                    if data['numero_compte'].startswith(elt):
                        list_code_efi.append(obj['ref'])
            list_code_efi = list(set(list_code_efi))
            sheet["E" + new_iteration] = ','.join(list_code_efi)

        namefile = "piste_audit.xlsx"
        wb.save(namefile)

    # ---------- Extract grouping Excel ----------
    def extract_grouping(self, id_mission):
        # Créer un fichier Excel pour l'export du grouping
        wb = openpyxl.Workbook()
        sheet = wb.active

        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        headers = ['Numéro compte', 'Solde n', 'Solde n-1', 'Grouping', 'Variation', 'Variation %', 'Compte qualitatif', 'Compte quantitatif', 'Compte significatif']

        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        balances = mission['balance_variation']
        grouping = mission['grouping']
        materiality = next(item for item in mission['materiality'] if item['choice'] is True)

        for i in range(len(columns)):
            sheet[columns[i] + '1'] = headers[i]

        for iteration, data in enumerate(balances):
            new_iteration = str(iteration + 2)
            sheet["A" + new_iteration] = data.get("numero_compte")
            sheet["B" + new_iteration] = data.get("solde_n")
            sheet["C" + new_iteration] = data.get("solde_n1")

            value_grouping = data.get("numero_compte")[0:2]
            variation = data.get("solde_n") - data.get("solde_n1")

            if variation == 0:
                variation_percent = 0
            elif data.get("solde_n1") == 0:
                variation_percent = 100
            else:
                variation_percent = (variation / data.get("solde_n1")) * 100

            sheet["D" + new_iteration] = value_grouping
            sheet["E" + new_iteration] = variation
            sheet["F" + new_iteration] = variation_percent
            sheet["G" + new_iteration] = next(item['significant'] for item in grouping if item['compte'] == value_grouping)
            sheet["H" + new_iteration] = next(item['materiality'] for item in grouping if item['compte'] == value_grouping)
            sheet["I" + new_iteration] = next(item['mat_sign'] for item in grouping if item['compte'] == value_grouping)

        second_sheet = wb.create_sheet(title="Seuil de matérialité")
        second_headers = ['materiality', 'performance materiality', 'trivial misstatements']
        second_sheet["A1"] = second_headers[0]
        second_sheet["B1"] = second_headers[1]
        second_sheet["C1"] = second_headers[2]

        second_sheet["A2"] = materiality['materiality']
        second_sheet["B2"] = materiality['performance_materiality']
        second_sheet["C2"] = materiality['trivial_misstatements']

        excel_io = BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)
        return excel_io

    # ==============================
    #  CONTROLES — Cohérence & Intangibilité
    # ==============================
    def _load_balance(self, balance_id):
        bal = db.Balance.find_one({"_id": ObjectId(balance_id)})
        return bal["balance"] if bal else []

    def _coherence_checks_for_year(self, lines):
        report = {"equilibre_global": True, "erreurs": []}

        # 1) Equilibre global
        sum_deb_fin = sum(int(x.get("debit_fin", 0) or 0) for x in lines)
        sum_cre_fin = sum(int(x.get("credit_fin", 0) or 0) for x in lines)
        if sum_deb_fin != sum_cre_fin:
            report["equilibre_global"] = False
            report["erreurs"].append({
                "type": "equilibre_global",
                "message": f"Total débit fin ({sum_deb_fin}) ≠ total crédit fin ({sum_cre_fin})"
            })

        # 2) Identité & signe
        for x in lines:
            di = int(x.get("debit_initial", 0) or 0)
            ci = int(x.get("credit_initial", 0) or 0)
            dm = int(x.get("debit_mvt", 0) or 0)
            cm = int(x.get("credit_mvt", 0) or 0)
            df = int(x.get("debit_fin", 0) or 0)
            cf = int(x.get("credit_fin", 0) or 0)

            lhs = df - cf
            rhs = (di - ci) + (dm - cm)
            if lhs != rhs:
                report["erreurs"].append({
                    "type": "identite_compte",
                    "numero_compte": x.get("numero_compte"),
                    "message": f"(df-cf)={lhs} ≠ (di-ci+dm-cm)={rhs}"
                })

            solde_reel = int(x.get("solde_reel", 0) or 0)
            sign = x.get("sign_solde")
            expect_sign = "D" if solde_reel >= 0 else "C"
            if sign not in ("D", "C") or sign != expect_sign:
                report["erreurs"].append({
                    "type": "signe_incoherent",
                    "numero_compte": x.get("numero_compte"),
                    "message": f"solde_reel={solde_reel} / sign_solde={sign} (attendu: {expect_sign})"
                })

        # ➕ Récap totaux pour UI
        report["totaux"] = {
            "debit_fin": sum_deb_fin,
            "credit_fin": sum_cre_fin,
            "nb_erreurs": len(report["erreurs"])
        }

        return report

    def controle_coherence(self, id_mission):
        print(f"=== DÉBUT contrôle_coherence pour mission {id_mission} ===")
        
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        print(f"Mission trouvée : {mission is not None}")
        
        if not mission:
            print(f"ERREUR : Mission {id_mission} non trouvée")
            return {"error": "Mission non trouvée"}
        
        bal_ids = mission.get("balances", [])
        print(f"Balances trouvées : {len(bal_ids)} - IDs: {bal_ids}")
        
        if not bal_ids:
            print("ERREUR : Aucune balance trouvée pour cette mission")
            return {"error": "Aucune balance trouvée"}
        
        out = {}

        # Récupérer l'année auditée de la mission
        annee_auditee = int(mission.get("annee_auditee", 0))

        for idx, bal_id in enumerate(bal_ids):
            # Calculer l'année de chaque balance
            annee_balance = annee_auditee - idx
            print(f"Traitement balance {idx} (année {annee_balance}) : {bal_id}")
            
            lines = self._load_balance(bal_id)
            print(f"Lignes de balance chargées : {len(lines)}")
            
            # Ajouter l'année au rapport
            rapport = self._coherence_checks_for_year(lines)
            rapport["annee"] = annee_balance
            
            out[str(annee_balance)] = rapport

        print(f"Rapport final généré : {out}")

        db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"controle_coherence": out}})
        print("=== FIN contrôle_coherence ===")
        return out

    def _index_by_compte(self, lines):
        return {str(x.get("numero_compte")): x for x in lines if x.get("numero_compte")}

    def controle_intangibilite(self, id_mission):
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})
        bal_ids = mission.get("balances", [])
        if len(bal_ids) < 2:
            return {"ok": False, "message": "Il faut au moins N et N-1."}

        bal_N = self._load_balance(bal_ids[0])
        bal_N1 = self._load_balance(bal_ids[1])
        idxN = self._index_by_compte(bal_N)
        idxN1 = self._index_by_compte(bal_N1)

        erreurs = []
        for num, ln in idxN.items():
            # classes 1..5 : bilan
            if not str(num).startswith(("1", "2", "3", "4", "5")):
                continue

            di = int(ln.get("debit_initial", 0) or 0)
            ci = int(ln.get("credit_initial", 0) or 0)
            ouvN = di - ci

            prev = idxN1.get(num)
            if prev:
                df = int(prev.get("debit_fin", 0) or 0)
                cf = int(prev.get("credit_fin", 0) or 0)
                clotN1 = df - cf
                if ouvN != clotN1:
                    ecart = ouvN - clotN1
                    erreurs.append({
                        "numero_compte": num,
                        "ouverture_n": ouvN,
                        "cloture_n1": clotN1,
                        "ecart": ecart,
                        "message": f"Ouverture N {ouvN} ≠ Clôture N-1 {clotN1}",
                        "justification": f"Écart de {ecart} entre l'ouverture de l'exercice N ({ouvN}) et la clôture de l'exercice N-1 ({clotN1}). Cette différence peut être due à des opérations de clôture, des reclassements comptables ou des corrections d'erreurs.",
                        "conclusion_audit": "Écart significatif détecté - Nécessite une justification et une documentation des causes de cette variation."
                    })
            else:
                erreurs.append({
                    "numero_compte": num,
                    "ouverture_n": ouvN,
                    "cloture_n1": None,
                    "ecart": None,
                    "message": "Compte présent en N mais absent en N-1",
                    "justification": f"Le compte {num} est présent dans l'exercice N avec un solde d'ouverture de {ouvN}, mais n'existait pas dans l'exercice N-1. Cela peut indiquer une création de compte, un reclassement ou une erreur de saisie.",
                    "conclusion_audit": "Compte nouvellement créé ou reclassé - Vérifier la légitimité de cette création et documenter les raisons."
                })

        report = {"ok": len(erreurs) == 0, "ecarts": erreurs}
        db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"controle_intangibilite": report}})
        return report


        return materiality



    # ---------- Analyses grouping ----------

    def make_quantitative_analysis(self, id_mission):

        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})

        grouping = mission['grouping']

        materiality = next((mat for mat in mission['materiality'] if mat.get('choice')), None)



        if materiality is not None:

            for item in grouping:

                value = False

                if int(item['solde_n']) >= int(materiality['materiality']):

                    value = True

                item['materiality'] = value



            result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})

            return result.modified_count

        else:

            return 0



    def make_qualitative_analysis(self, id_mission, significant):

        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})

        grouping = mission['grouping']



        for item in grouping:

            value = next((elt['significant'] for elt in significant if elt.get('compte') == item.get('compte')), None)

            if value is None:

                item['significant'] = False

            else:

                item['significant'] = value



        result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})

        return result.modified_count



    def make_final_sm(self, id_mission):

        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})

        grouping = mission['grouping']



        for item in grouping:

            mat = item.get('materiality', None)

            sign = item.get('significant', None)



            if (mat is not None) and (sign is not None):

                value = ""

                if mat is False and sign is True:

                    value = "non matériel significatif"

                elif mat is False and sign is False:

                    value = "non matériel non significatif"

                elif mat is True and sign is True:

                    value = "matériel significatif"

                elif mat is True and sign is False:

                    value = "matériel non significatif"

                else:

                    value = None

                item['mat_sign'] = value



        result = db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"grouping": grouping}})

        return result.modified_count, grouping



    # ---------- Récup mission ----------

    def afficher_informations_missions(self, id_client):

        _id = id_client

        query = db.Mission1.find_one({"_id": _id})

        query['_id'] = str(query['_id'])

        return query



    # ---------- Production COMMENTAIRE ----------

    def prod_efi(self, balance_n, balance_n1, balance_variation):

        mapping_path = os.path.join(os.path.dirname(__file__), "..", "mapping_efi.json")

        with open(mapping_path, 'r', encoding='utf-8') as file:

            result = json.load(file)

        mapping = result['structure']



        # nettoyer les champs *_cpt en listes

        for mapp in mapping:

            mapp['brut_cpt'] = mapp['brut_cpt'].split(',') if mapp.get('brut_cpt') is not None else mapp.get('brut_cpt')

            mapp['amor_cpt'] = mapp['amor_cpt'].split(',') if mapp.get('amor_cpt') is not None else mapp.get('amor_cpt')

            mapp['net_cpt'] = mapp['net_cpt'].split(',') if mapp.get('net_cpt') is not None else mapp.get('net_cpt')

            mapp['brut_except_cpt'] = mapp['brut_except_cpt'].split(',') if mapp.get('brut_except_cpt') is not None else mapp.get('brut_except_cpt', [])

            mapp['amor_except_cpt'] = mapp['amor_except_cpt'].split(',') if mapp.get('amor_except_cpt') is not None else mapp.get('amor_except_cpt', [])

            mapp['net_except_cpt'] = mapp['net_except_cpt'].split(',') if mapp.get('net_except_cpt') is not None else mapp.get('net_except_cpt', [])



        datum = {}

        list_efi = ['actif', 'passif', 'pnl']

        for efi in list_efi:

            structure = []

            select_mapping = (elt for elt in mapping if elt['nature'] == efi)

            for data in select_mapping:

                row = {}

                if data.get('brut_cpt') and data.get('amor_cpt'):

                    # N

                    brut_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['brut_cpt']))

                    amor_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['amor_cpt']))



                    brut_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('brut_except_cpt', [])))

                    amor_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('amor_except_cpt', [])))



                    data['brut_solde_n'] = brut_solde_n + brut_except_n

                    data['amor_solde_n'] = amor_solde_n + amor_except_n

                    data['net_solde_n'] = data['brut_solde_n'] + data['amor_solde_n']



                    # N-1

                    brut_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['brut_cpt']))

                    amor_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['amor_cpt']))

                    net_except_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))

                    data['net_solde_n1'] = brut_n1 + amor_n1 + net_except_n1

                else:

                    net_solde_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data['net_cpt']))

                    net_solde_n1 = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data['net_cpt']))



                    net_except_n = sum(item['solde_reel'] for item in balance_n if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))

                    net_except_n1_bis = sum(item['solde_reel'] for item in balance_n1 if any(str(item['numero_compte']).startswith(cpt) for cpt in data.get('net_except_cpt', [])))



                    data['net_solde_n'] = net_solde_n + net_except_n

                    data['net_solde_n1'] = net_solde_n1 + net_except_n1_bis



                row['ref'] = data['ref']

                row['libelle'] = data['libelle']

                row['compte_to_be_used'] = str(data.get('brut_cpt')) + str(data.get('amor_cpt')) + str(data.get('net_cpt')) + str(data.get('brut_except_cpt')) + str(data.get('amor_except_cpt')) + str(data.get('net_except_cpt'))

                row['compte_to_be_used'] = row['compte_to_be_used'].replace('None', '')



                one = data.get('brut_cpt', []) or []

                two = data.get('amor_cpt', []) or []

                three = data.get('net_cpt', []) or []

                four = data.get('brut_except_cpt', []) or []

                five = data.get('amor_except_cpt', []) or []

                six = data.get('net_except_cpt', []) or []



                row['compte_to_be_used_off'] = list(set(one + two + three + four + five + six))



                row['brut_solde_n'] = data.get('brut_solde_n')

                row['amor_solde_n'] = data.get('amor_solde_n')

                row['net_solde_n'] = data.get('net_solde_n')

                row['net_solde_n1'] = data.get('net_solde_n1')



                structure.append(row)



            datum[efi] = structure



        return datum



    # ---------- Piste d'audit ----------

    def audit_trail(self, id_mission):

        # Créer un fichier Excel pour la piste d'audit
        wb = openpyxl.Workbook()

        sheet = wb.active



        columns = ['A', 'B', 'C', 'D', 'E']

        headers = ['Numéro compte', 'Solde n', 'Solde n-1', 'Grouping', 'Code COMMENTAIRE']



        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})

        balances = mission['balance_variation']

        actif = mission['efi']['actif']

        passif = mission['efi']['passif']

        pnl = mission['efi']['pnl']



        efi = actif + passif + pnl



        for i in range(len(columns)):

            sheet[columns[i] + '1'] = headers[i]



        for iteration, data in enumerate(balances):

            new_iteration = str(iteration + 2)

            sheet["A" + new_iteration] = data.get("numero_compte")

            sheet["B" + new_iteration] = data.get("solde_n")

            sheet["C" + new_iteration] = data.get("solde_n1")

            sheet["D" + new_iteration] = data.get("numero_compte")[0:2]



            list_code_efi = []

            for obj in efi:

                for elt in obj['compte_to_be_used_off']:

                    if data['numero_compte'].startswith(elt):

                        list_code_efi.append(obj['ref'])

            list_code_efi = list(set(list_code_efi))

            sheet["E" + new_iteration] = ','.join(list_code_efi)



        namefile = "piste_audit.xlsx"

        wb.save(namefile)



    # ---------- Extract grouping Excel ----------

    def extract_grouping(self, id_mission):

        # Créer un fichier Excel pour l'export du grouping
        wb = openpyxl.Workbook()

        sheet = wb.active



        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

        headers = ['Numéro compte', 'Solde n', 'Solde n-1', 'Grouping', 'Variation', 'Variation %', 'Compte qualitatif', 'Compte quantitatif', 'Compte significatif']



        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})

        balances = mission['balance_variation']

        grouping = mission['grouping']

        materiality = next(item for item in mission['materiality'] if item['choice'] is True)



        for i in range(len(columns)):

            sheet[columns[i] + '1'] = headers[i]



        for iteration, data in enumerate(balances):

            new_iteration = str(iteration + 2)

            sheet["A" + new_iteration] = data.get("numero_compte")

            sheet["B" + new_iteration] = data.get("solde_n")

            sheet["C" + new_iteration] = data.get("solde_n1")



            value_grouping = data.get("numero_compte")[0:2]

            variation = data.get("solde_n") - data.get("solde_n1")



            if variation == 0:

                variation_percent = 0

            elif data.get("solde_n1") == 0:

                variation_percent = 100

            else:

                variation_percent = (variation / data.get("solde_n1")) * 100



            sheet["D" + new_iteration] = value_grouping

            sheet["E" + new_iteration] = variation

            sheet["F" + new_iteration] = variation_percent

            sheet["G" + new_iteration] = next(item['significant'] for item in grouping if item['compte'] == value_grouping)

            sheet["H" + new_iteration] = next(item['materiality'] for item in grouping if item['compte'] == value_grouping)

            sheet["I" + new_iteration] = next(item['mat_sign'] for item in grouping if item['compte'] == value_grouping)



        second_sheet = wb.create_sheet(title="Seuil de matérialité")

        second_headers = ['materiality', 'performance materiality', 'trivial misstatements']

        second_sheet["A1"] = second_headers[0]

        second_sheet["B1"] = second_headers[1]

        second_sheet["C1"] = second_headers[2]



        second_sheet["A2"] = materiality['materiality']

        second_sheet["B2"] = materiality['performance_materiality']

        second_sheet["C2"] = materiality['trivial_misstatements']



        excel_io = BytesIO()

        wb.save(excel_io)

        excel_io.seek(0)

        return excel_io



    # ==============================

    #  CONTROLES — Cohérence & Intangibilité

    # ==============================

    def _load_balance(self, balance_id):

        bal = db.Balance.find_one({"_id": ObjectId(balance_id)})

        return bal["balance"] if bal else []



    def _coherence_checks_for_year(self, lines):

        report = {"equilibre_global": True, "erreurs": []}



        # 1) Equilibre global

        sum_deb_fin = sum(int(x.get("debit_fin", 0) or 0) for x in lines)

        sum_cre_fin = sum(int(x.get("credit_fin", 0) or 0) for x in lines)

        if sum_deb_fin != sum_cre_fin:

            report["equilibre_global"] = False

            report["erreurs"].append({

                "type": "equilibre_global",

                "message": f"Total débit fin ({sum_deb_fin}) ≠ total crédit fin ({sum_cre_fin})"

            })



        # 2) Identité & signe

        for x in lines:

            di = int(x.get("debit_initial", 0) or 0)

            ci = int(x.get("credit_initial", 0) or 0)

            dm = int(x.get("debit_mvt", 0) or 0)

            cm = int(x.get("credit_mvt", 0) or 0)

            df = int(x.get("debit_fin", 0) or 0)

            cf = int(x.get("credit_fin", 0) or 0)



            lhs = df - cf

            rhs = (di - ci) + (dm - cm)

            if lhs != rhs:

                report["erreurs"].append({

                    "type": "identite_compte",

                    "numero_compte": x.get("numero_compte"),

                    "message": f"(df-cf)={lhs} ≠ (di-ci+dm-cm)={rhs}"

                })



            solde_reel = int(x.get("solde_reel", 0) or 0)

            sign = x.get("sign_solde")

            expect_sign = "D" if solde_reel >= 0 else "C"

            if sign not in ("D", "C") or sign != expect_sign:

                report["erreurs"].append({

                    "type": "signe_incoherent",

                    "numero_compte": x.get("numero_compte"),

                    "message": f"solde_reel={solde_reel} / sign_solde={sign} (attendu: {expect_sign})"

                })



        # ➕ Récap totaux pour UI

        report["totaux"] = {

            "debit_fin": sum_deb_fin,

            "credit_fin": sum_cre_fin,

            "nb_erreurs": len(report["erreurs"])

        }



        return report



    def controle_coherence(self, id_mission):

        print(f"=== DÉBUT contrôle_coherence pour mission {id_mission} ===")
        
        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})

        print(f"Mission trouvée : {mission is not None}")
        
        if not mission:
            print(f"ERREUR : Mission {id_mission} non trouvée")
            return {"error": "Mission non trouvée"}
        
        bal_ids = mission.get("balances", [])

        print(f"Balances trouvées : {len(bal_ids)} - IDs: {bal_ids}")
        
        if not bal_ids:
            print("ERREUR : Aucune balance trouvée pour cette mission")
            return {"error": "Aucune balance trouvée"}
        
        out = {}


        # Récupérer l'année auditée de la mission
        annee_auditee = int(mission.get("annee_auditee", 0))


        for idx, bal_id in enumerate(bal_ids):

            # Calculer l'année de chaque balance
            annee_balance = annee_auditee - idx
            print(f"Traitement balance {idx} (année {annee_balance}) : {bal_id}")
            
            lines = self._load_balance(bal_id)

            print(f"Lignes de balance chargées : {len(lines)}")
            
            # Ajouter l'année au rapport
            rapport = self._coherence_checks_for_year(lines)
            rapport["annee"] = annee_balance
            
            out[str(annee_balance)] = rapport

        print(f"Rapport final généré : {out}")


        db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"controle_coherence": out}})

        print("=== FIN contrôle_coherence ===")
        return out



    def _index_by_compte(self, lines):

        return {str(x.get("numero_compte")): x for x in lines if x.get("numero_compte")}



    def controle_intangibilite(self, id_mission):

        mission = db.Mission1.find_one({"_id": ObjectId(id_mission)})

        bal_ids = mission.get("balances", [])

        if len(bal_ids) < 2:

            return {"ok": False, "message": "Il faut au moins N et N-1."}



        bal_N = self._load_balance(bal_ids[0])

        bal_N1 = self._load_balance(bal_ids[1])

        idxN = self._index_by_compte(bal_N)

        idxN1 = self._index_by_compte(bal_N1)



        erreurs = []

        for num, ln in idxN.items():

            # classes 1..5 : bilan

            if not str(num).startswith(("1", "2", "3", "4", "5")):

                continue



            di = int(ln.get("debit_initial", 0) or 0)

            ci = int(ln.get("credit_initial", 0) or 0)

            ouvN = di - ci



            prev = idxN1.get(num)

            if prev:

                df = int(prev.get("debit_fin", 0) or 0)

                cf = int(prev.get("credit_fin", 0) or 0)

                clotN1 = df - cf

                if ouvN != clotN1:

                    ecart = ouvN - clotN1
                    erreurs.append({

                        "numero_compte": num,

                        "ouverture_n": ouvN,

                        "cloture_n1": clotN1,

                        "ecart": ecart,
                        "message": f"Ouverture N {ouvN} ≠ Clôture N-1 {clotN1}",
                        "justification": f"Écart de {ecart} entre l'ouverture de l'exercice N ({ouvN}) et la clôture de l'exercice N-1 ({clotN1}). Cette différence peut être due à des opérations de clôture, des reclassements comptables ou des corrections d'erreurs.",
                        "conclusion_audit": "Écart significatif détecté - Nécessite une justification et une documentation des causes de cette variation."
                    })

            else:

                erreurs.append({

                    "numero_compte": num,

                    "ouverture_n": ouvN,

                    "cloture_n1": None,

                    "ecart": None,

                    "message": "Compte présent en N mais absent en N-1",
                    "justification": f"Le compte {num} est présent dans l'exercice N avec un solde d'ouverture de {ouvN}, mais n'existait pas dans l'exercice N-1. Cela peut indiquer une création de compte, un reclassement ou une erreur de saisie.",
                    "conclusion_audit": "Compte nouvellement créé ou reclassé - Vérifier la légitimité de cette création et documenter les raisons."
                })



        report = {"ok": len(erreurs) == 0, "ecarts": erreurs}

        db.Mission1.update_one({"_id": ObjectId(id_mission)}, {"$set": {"controle_intangibilite": report}})

        return report




